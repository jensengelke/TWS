#!/usr/bin/env python3
"""
Interactive Brokers Flex Query zu Excel Converter
Konvertiert IB Flex Query XML zu strukturierter Excel-Datei mit Kauf/Verkauf-Paaren
"""

import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import requests
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
import argparse
import sys

class IBFlexQueryProcessor:
    def __init__(self):
        self.trades_df = None
        self.matched_trades = []
        self.unmatched_trades = []
        
    def download_flex_query(self, token, query_id):
        """
        Lädt Flex Query direkt von IB herunter
        """
        print(f"Lade Flex Query {query_id} herunter...")
        
        # Schritt 1: Query ausführen
        request_url = f"https://gdcdyn.interactivebrokers.com/Universal/servlet/FlexStatementService.SendRequest"
        params = {
            't': token,
            'q': query_id,
            'v': '3'
        }
        
        response = requests.get(request_url, params=params)
        if response.status_code != 200:
            raise Exception(f"Fehler beim Ausführen der Query: {response.status_code}")
            
        # Parse Response für Reference Code
        root = ET.fromstring(response.text)
        if root.find('.//Status').text != 'Success':
            error_msg = root.find('.//ErrorMessage')
            raise Exception(f"Query Fehler: {error_msg.text if error_msg is not None else 'Unbekannter Fehler'}")
            
        reference_code = root.find('.//ReferenceCode').text
        print(f"Reference Code erhalten: {reference_code}")
        
        # Schritt 2: Daten abrufen
        download_url = f"https://gdcdyn.interactivebrokers.com/Universal/servlet/FlexStatementService.GetStatement"
        params = {
            't': token,
            'q': reference_code,
            'v': '3'
        }
        
        response = requests.get(download_url, params=params)
        if response.status_code != 200:
            raise Exception(f"Fehler beim Herunterladen: {response.status_code}")
            
        return response.text
    
    def parse_xml_file(self, xml_file_path):
        """
        Parst XML-Datei von lokaler Datei
        """
        print(f"Parse XML-Datei: {xml_file_path}")
        tree = ET.parse(xml_file_path)
        return tree.getroot()
    
    def parse_xml_content(self, xml_content):
        """
        Parst XML-Content als String
        """
        print("Parse XML-Content...")
        return ET.fromstring(xml_content)
    
    def extract_trades(self, root):
        """
        Extrahiert Trade-Daten aus XML
        """
        trades = []
        
        # Suche nach Trade-Elementen in verschiedenen möglichen Strukturen
        trade_elements = (
            root.findall('.//Trade') + 
            root.findall('.//Trades/Trade') +
            root.findall('.//FlexStatement/Trades/Trade')
        )
        
        if not trade_elements:
            print("Warnung: Keine Trade-Elemente gefunden!")
            return pd.DataFrame()
        
        print(f"Gefunden: {len(trade_elements)} Trades")
        
        for trade in trade_elements:
            trade_data = {
                'symbol': trade.get('symbol', ''),
                'description': trade.get('description', ''),
                'isin': trade.get('isin', ''),
                'date': trade.get('dateTime', trade.get('tradeDate', '')),
                'settlementDate': trade.get('settleDate', ''),
                'buySell': trade.get('buySell', ''),
                'quantity': float(trade.get('quantity', 0)),
                'price': float(trade.get('tradePrice', 0)),
                'amount': float(trade.get('proceeds', 0)),
                'commission': float(trade.get('ibCommission', 0)),
                'fees': float(trade.get('fees', 0)),
                'netCash': float(trade.get('netCash', 0)),
                'currency': trade.get('currency', ''),
                'accountId': trade.get('accountId', ''),
                'orderTime': trade.get('orderTime', ''),
                'exchange': trade.get('exchange', ''),
                'multiplier': float(trade.get('multiplier', 1)),
                'assetCategory': trade.get('assetCategory', ''),
                'subCategory': trade.get('subCategory', '')
            }
            trades.append(trade_data)
        
        df = pd.DataFrame(trades)
        
        # Datum konvertieren
        if not df.empty and 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            if 'settlementDate' in df.columns:
                df['settlementDate'] = pd.to_datetime(df['settlementDate'], errors='coerce')
        
        print(f"Trades erfolgreich extrahiert: {len(df)} Einträge")
        return df
    
    def match_buy_sell_pairs(self, df):
        """
        Ordnet Kauf- und Verkauf-Transaktionen zu
        """
        if df.empty:
            return [], []
        
        # Gruppiere nach Symbol
        grouped = df.groupby('symbol')
        matched_pairs = []
        unmatched = []
        
        for symbol, group in grouped:
            print(f"Verarbeite Symbol: {symbol}")
            
            # Sortiere nach Datum
            group = group.sort_values('date')
            
            # Separiere Käufe und Verkäufe
            buys = group[group['buySell'] == 'BUY'].copy()
            sells = group[group['buySell'] == 'SELL'].copy()
            
            # FIFO-Matching
            buy_queue = buys.to_dict('records')
            
            for sell_row in sells.to_dict('records'):
                sell_qty = abs(sell_row['quantity'])
                original_sell_qty = sell_qty  # Für Short-Position Tracking
                
                while sell_qty > 0 and buy_queue:
                    buy_row = buy_queue[0]
                    buy_qty = buy_row['quantity']
                    
                    # Bestimme übereinstimmende Menge
                    match_qty = min(buy_qty, sell_qty)
                    
                    # Erstelle Match
                    matched_pair = {
                        'symbol': symbol,
                        'description': buy_row['description'],
                        'currency': buy_row['currency'],
                        'quantity': match_qty,
                        
                        # Kauf-Daten (mit Multiplier)
                        'buy_date': buy_row['date'],
                        'buy_price': buy_row['price'],
                        'buy_amount': buy_row['price'] * match_qty * buy_row['multiplier'],
                        'buy_commission': (buy_row['commission'] / buy_row['quantity']) * match_qty if buy_row['quantity'] != 0 else 0,
                        'buy_multiplier': buy_row['multiplier'],
                        
                        # Verkauf-Daten (mit Multiplier)
                        'sell_date': sell_row['date'],
                        'sell_price': sell_row['price'],
                        'sell_amount': sell_row['price'] * match_qty * sell_row['multiplier'],
                        'sell_commission': (abs(sell_row['commission']) / abs(sell_row['quantity'])) * match_qty if sell_row['quantity'] != 0 else 0,
                        
                        # Berechnungen (mit korrekten Beträgen)
                        'holding_period_days': (sell_row['date'] - buy_row['date']).days if pd.notnull(sell_row['date']) and pd.notnull(buy_row['date']) else 0,
                        'total_commission': 0,
                        'gross_pnl': 0,
                        'net_pnl': 0,
                        'return_percent': 0
                    }
                    
                    # Berechne P&L mit korrekten Beträgen (nur Kommissionen, keine separaten Gebühren)
                    matched_pair['total_commission'] = matched_pair['buy_commission'] + matched_pair['sell_commission']
                    matched_pair['gross_pnl'] = matched_pair['sell_amount'] - matched_pair['buy_amount']
                    matched_pair['net_pnl'] = matched_pair['gross_pnl'] - matched_pair['total_commission']
                    
                    # Berechne prozentuale Rendite
                    matched_pair['return_percent'] = 0
                    if matched_pair['buy_amount'] != 0:
                        matched_pair['return_percent'] = (matched_pair['net_pnl'] / matched_pair['buy_amount']) * 100
                    
                    # Formatiere Währungs- und Prozentangaben
                    currency = buy_row.get('currency', 'USD')
                    matched_pair['net_pnl_formatted'] = f"{matched_pair['net_pnl']:.2f} {currency}"
                    matched_pair['return_percent_formatted'] = f"{matched_pair['return_percent']:.1f}%"
                    matched_pair['gross_pnl_formatted'] = f"{matched_pair['gross_pnl']:.2f} {currency}"
                    matched_pair['buy_amount_formatted'] = f"{matched_pair['buy_amount']:.2f} {currency}"
                    matched_pair['sell_amount_formatted'] = f"{matched_pair['sell_amount']:.2f} {currency}"
                    
                    matched_pairs.append(matched_pair)
                    
                    # Update Mengen
                    sell_qty -= match_qty
                    buy_row['quantity'] -= match_qty
                    
                    if buy_row['quantity'] <= 0:
                        buy_queue.pop(0)
                
                # Nicht zugeordnete Verkäufe (Short-Positionen)
                if sell_qty > 0:
                    unmatched.append({
                        'type': 'SELL (Short)',
                        'symbol': symbol,
                        'date': sell_row['date'],
                        'quantity': -sell_qty,  # Negativ für Short
                        'price': sell_row['price'],
                        'amount': sell_row['price'] * sell_qty * sell_row['multiplier'],
                        'multiplier': sell_row['multiplier'],
                        'currency': sell_row['currency']
                    })
            
            # Nicht zugeordnete Käufe
            for buy_row in buy_queue:
                if buy_row['quantity'] > 0:
                    unmatched.append({
                        'type': 'BUY',
                        'symbol': symbol,
                        'date': buy_row['date'],
                        'quantity': buy_row['quantity'],
                        'price': buy_row['price'],
                        'amount': buy_row['price'] * buy_row['quantity'] * buy_row['multiplier'],
                        'multiplier': buy_row['multiplier'],
                        'currency': buy_row['currency']
                    })
        
        print(f"Matches gefunden: {len(matched_pairs)}")
        print(f"Unmatched Trades: {len(unmatched)}")
        
        return matched_pairs, unmatched
    
    def create_excel_report(self, matched_pairs, unmatched, output_file):
        """
        Erstellt Excel-Report mit Formatierung
        """
        print(f"Erstelle Excel-Report: {output_file}")
        
        # Stelle sicher, dass der Output-Dateiname .xlsx Endung hat
        if not output_file.lower().endswith('.xlsx'):
            output_file = output_file.rsplit('.', 1)[0] + '.xlsx'
        
        try:
            # Verwende explizit openpyxl engine
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                
                # Sheet 1: Matched Trades
                if matched_pairs:
                    matched_df = pd.DataFrame(matched_pairs)
                    
                    # Reorganisiere Spalten für bessere Übersicht (ohne unwanted columns)
                    column_order = [
                        'symbol', 'description', 'currency', 'quantity',
                        'buy_date', 'buy_price', 'buy_amount_formatted',
                        'sell_date', 'sell_price', 'sell_amount_formatted',
                        'holding_period_days', 
                        'gross_pnl_formatted', 'net_pnl_formatted', 'return_percent_formatted',
                        'buy_multiplier'
                    ]
                    
                    # Nur verfügbare Spalten verwenden
                    available_columns = [col for col in column_order if col in matched_df.columns]
                    if available_columns:
                        matched_df = matched_df[available_columns]
                    
                    matched_df.to_excel(writer, sheet_name='Matched_Trades', index=False)
                    
                    # Formatierung für Matched Trades
                    ws = writer.sheets['Matched_Trades']
                    self._format_worksheet(ws, matched_df)
                
                # Sheet 2: Unmatched Trades (Offene Positionen)
                if unmatched:
                    unmatched_df = pd.DataFrame(unmatched)
                    
                    # Formatiere auch unmatched trades
                    for idx, row in unmatched_df.iterrows():
                        currency = row.get('currency', 'USD')
                        unmatched_df.at[idx, 'amount_formatted'] = f"{row['amount']:.2f} {currency}"
                        
                        # Berechne aktuellen unrealisierten Gewinn/Verlust (vereinfacht)
                        # Hier könnten Sie aktuelle Marktpreise einbinden
                        unmatched_df.at[idx, 'position_value'] = f"{abs(row['quantity']) * row['price'] * row['multiplier']:.2f} {currency}"
                        unmatched_df.at[idx, 'position_type'] = 'Long Position' if row['type'] == 'BUY' else 'Short Position'
                    
                    # Reorganisiere Spalten (ohne ungewollte Felder)
                    unmatched_columns = ['symbol', 'position_type', 'date', 'quantity', 'price', 'position_value', 'currency']
                    available_unmatched_cols = [col for col in unmatched_columns if col in unmatched_df.columns]
                    if available_unmatched_cols:
                        unmatched_df = unmatched_df[available_unmatched_cols]
                    
                    unmatched_df.to_excel(writer, sheet_name='Open_Positions', index=False)
                    
                    ws = writer.sheets['Open_Positions']
                    self._format_worksheet(ws, unmatched_df, is_unmatched=True)
                
                # Sheet 3: Zusammenfassung
                if matched_pairs:
                    summary = self._create_summary(matched_pairs, unmatched)
                    summary.to_excel(writer, sheet_name='Summary', index=False)
            
            print(f"Excel-Report erstellt: {output_file}")
            
        except Exception as e:
            print(f"Fehler beim Excel-Export: {e}")
            # Fallback: Einfache CSV-Exports
            print("Erstelle CSV-Fallback...")
            
            if matched_pairs:
                matched_df = pd.DataFrame(matched_pairs)
                csv_file = output_file.replace('.xlsx', '_matched.csv')
                matched_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                print(f"Matched Trades CSV: {csv_file}")
            
            if unmatched:
                unmatched_df = pd.DataFrame(unmatched)
                csv_file = output_file.replace('.xlsx', '_unmatched.csv')
                unmatched_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                print(f"Unmatched Trades CSV: {csv_file}")
                
            raise
    
    def _format_worksheet(self, ws, df, is_unmatched=False):
        """
        Formatiert Excel-Arbeitsblatt
        """
        # Header-Formatierung
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Spaltenbreiten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Bedingte Formatierung für P&L (nur bei matched trades)
        if not is_unmatched and any('net_pnl' in col for col in df.columns):
            try:
                # Grün für Gewinn, Rot für Verlust
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Finde P&L Spalte (formatierte Version)
                pnl_col = None
                for idx, col in enumerate(df.columns):
                    if 'net_pnl_formatted' in col:
                        pnl_col = idx + 1
                        break
                
                # Falls formatierte Version nicht gefunden, suche normale Version
                if not pnl_col:
                    for idx, col in enumerate(df.columns):
                        if 'net_pnl' in col and 'formatted' not in col:
                            pnl_col = idx + 1
                            break
                
                if pnl_col:
                    # Für formatierte Währungsbeträge müssen wir die Regeln anders setzen
                    # Da die Werte als Text vorliegen, verwenden wir eine einfachere Regel
                    pass  # Bedingte Formatierung für Textwerte ist komplexer, lassen wir vorerst weg
                    
            except Exception as e:
                print(f"Warnung: Konnte bedingte Formatierung nicht anwenden: {e}")
    
    def _create_summary(self, matched_pairs, unmatched=None):
        """
        Erstellt Zusammenfassung der Trades
        """
        df = pd.DataFrame(matched_pairs)
        
        summary_data = []
        
        # Gesamt-Statistiken (ohne separate Gebühren)
        total_trades = len(matched_pairs)
        total_gross_pnl = df['gross_pnl'].sum()
        total_net_pnl = df['net_pnl'].sum()
        total_commission = df['total_commission'].sum()
        winning_trades = len(df[df['net_pnl'] > 0])
        losing_trades = len(df[df['net_pnl'] < 0])
        
        # Gesamtinvestment (Kaufbeträge)
        total_investment = df['buy_amount'].sum()
        
        # Bestimme Hauptwährung (häufigste Währung)
        main_currency = df['currency'].mode().iloc[0] if not df.empty and 'currency' in df.columns else 'USD'
        
        summary_data.extend([
            {'Metrik': 'ABGESCHLOSSENE TRADES', 'Wert': ''},
            {'Metrik': 'Gesamt Trades', 'Wert': total_trades},
            {'Metrik': 'Gewinn-Trades', 'Wert': winning_trades},
            {'Metrik': 'Verlust-Trades', 'Wert': losing_trades},
            {'Metrik': 'Win Rate', 'Wert': f"{(winning_trades/total_trades)*100:.1f}%" if total_trades > 0 else "0.0%"},
            {'Metrik': 'Gesamtinvestment', 'Wert': f"{total_investment:.2f} {main_currency}"},
            {'Metrik': 'Gesamt Brutto P&L', 'Wert': f"{total_gross_pnl:.2f} {main_currency}"},
            {'Metrik': 'Gesamt Netto P&L', 'Wert': f"{total_net_pnl:.2f} {main_currency}"},
            {'Metrik': 'Gesamtrendite', 'Wert': f"{(total_net_pnl/total_investment)*100:.1f}%" if total_investment != 0 else "0.0%"},
            {'Metrik': 'Gesamt Kommissionen', 'Wert': f"{total_commission:.2f} {main_currency}"},
        ])
        
        # Offene Positionen
        if unmatched:
            unmatched_df = pd.DataFrame(unmatched)
            
            long_positions = unmatched_df[unmatched_df['type'] == 'BUY']
            short_positions = unmatched_df[unmatched_df['type'].str.contains('Short', na=False)]
            
            total_open_value = unmatched_df['amount'].sum()
            
            summary_data.extend([
                {'Metrik': '', 'Wert': ''},
                {'Metrik': 'OFFENE POSITIONEN', 'Wert': ''},
                {'Metrik': 'Long Positionen', 'Wert': len(long_positions)},
                {'Metrik': 'Short Positionen', 'Wert': len(short_positions)},
                {'Metrik': 'Gesamtwert offene Pos.', 'Wert': f"{total_open_value:.2f} {main_currency}"},
            ])
            
            # Details zu offenen Positionen
            for _, pos in unmatched_df.iterrows():
                pos_type = "Long" if pos['type'] == 'BUY' else "Short"
                pos_currency = pos.get('currency', main_currency)
                summary_data.append({
                    'Metrik': f'{pos["symbol"]} ({pos_type})', 
                    'Wert': f'{pos["quantity"]:.0f} @ {pos["price"]:.2f} = {pos["amount"]:.2f} {pos_currency}'
                })
        
        # Multiplier-Analyse
        if not df.empty:
            multiplier_info = df.groupby(['symbol', 'buy_multiplier']).size().reset_index(name='count')
            if not multiplier_info.empty:
                summary_data.extend([
                    {'Metrik': '', 'Wert': ''},  # Leerzeile
                    {'Metrik': 'MULTIPLIER INFO', 'Wert': ''},
                ])
                
                for _, row in multiplier_info.iterrows():
                    summary_data.append({
                        'Metrik': f'{row["symbol"]} (Multiplier: {row["buy_multiplier"]})', 
                        'Wert': f'{row["count"]} Trades'
                    })
        
        # Top Gewinner/Verlierer
        if not df.empty:
            best_trade = df.loc[df['net_pnl'].idxmax()]
            worst_trade = df.loc[df['net_pnl'].idxmin()]
            
            best_currency = best_trade.get('currency', main_currency)
            worst_currency = worst_trade.get('currency', main_currency)
            
            summary_data.extend([
                {'Metrik': '', 'Wert': ''},  # Leerzeile
                {'Metrik': 'TOP PERFORMER', 'Wert': ''},
                {'Metrik': f'Bester Trade ({best_trade["symbol"]})', 'Wert': f"{best_trade['net_pnl']:.2f} {best_currency} ({best_trade['return_percent']:.1f}%)"},
                {'Metrik': f'Schlechtester Trade ({worst_trade["symbol"]})', 'Wert': f"{worst_trade['net_pnl']:.2f} {worst_currency} ({worst_trade['return_percent']:.1f}%)"},
            ])
        
        return pd.DataFrame(summary_data)
    
    def process_flex_query(self, input_source, output_file, token=None, query_id=None):
        """
        Hauptverarbeitungsfunktion
        """
        try:
            # Bestimme Input-Quelle
            if token and query_id:
                xml_content = self.download_flex_query(token, query_id)
                root = self.parse_xml_content(xml_content)
            else:
                root = self.parse_xml_file(input_source)
            
            # Extrahiere Trades
            trades_df = self.extract_trades(root)
            
            if trades_df.empty:
                print("Keine Trades gefunden!")
                return
            
            # Zeige erste paar Trades zur Kontrolle
            print("\nErste 3 Trades:")
            display_cols = ['symbol', 'date', 'buySell', 'quantity', 'price', 'multiplier', 'amount']
            available_display_cols = [col for col in display_cols if col in trades_df.columns]
            print(trades_df[available_display_cols].head(3))
            
            # Zeige Multiplier-Info
            if not trades_df.empty and 'multiplier' in trades_df.columns:
                multiplier_summary = trades_df.groupby(['symbol', 'multiplier']).size().reset_index(name='count')
                print(f"\nMultiplier-Übersicht:")
                for _, row in multiplier_summary.iterrows():
                    print(f"  {row['symbol']}: Multiplier {row['multiplier']} ({row['count']} Trades)")
            
            # Matche Kauf/Verkauf Paare
            matched_pairs, unmatched = self.match_buy_sell_pairs(trades_df)
            
            # Erstelle Excel Report
            self.create_excel_report(matched_pairs, unmatched, output_file)
            
            print(f"\n✅ Verarbeitung abgeschlossen!")
            print(f"📊 Matched Trades: {len(matched_pairs)}")
            print(f"❓ Unmatched Trades: {len(unmatched)}")
            print(f"📁 Output: {output_file}")
            
        except Exception as e:
            print(f"❌ Fehler: {str(e)}")
            return False
        
        return True

def main():
    parser = argparse.ArgumentParser(description='IB Flex Query zu Excel Converter')
    parser.add_argument('--xml-file', '-f', help='Pfad zur XML-Datei')
    parser.add_argument('--token', '-t', help='IB Flex Query Token')
    parser.add_argument('--query-id', '-q', help='IB Flex Query ID')
    parser.add_argument('--output', '-o', default='ib_trades_analysis.xlsx', help='Output Excel-Datei')
    
    args = parser.parse_args()
    
    # Debug-Ausgabe
    print(f"Debug - Input file: {args.xml_file}")
    print(f"Debug - Output file: {args.output}")
    print(f"Debug - Pandas version: {pd.__version__}")
    
    processor = IBFlexQueryProcessor()
    
    # Validierung der Eingaben
    if not args.xml_file and not (args.token and args.query_id):
        print("❌ Bitte entweder XML-Datei oder Token+Query-ID angeben!")
        print("\nBeispiele:")
        print("  python ib_flex_processor.py -f flex_query.xml")
        print("  python ib_flex_processor.py -t YOUR_TOKEN -q YOUR_QUERY_ID")
        sys.exit(1)
    
    # Verarbeitung
    success = processor.process_flex_query(
        input_source=args.xml_file,
        output_file=args.output,
        token=args.token,
        query_id=args.query_id
    )
    
    if success:
        print("\n🎉 Erfolgreich abgeschlossen!")
    else:
        sys.exit(1)

if __name__ == "__main__":
    main()
